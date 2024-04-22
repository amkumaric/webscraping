import openpyxl as xl

wb = xl.load_workbook('example.xlsx')

sn = wb.sheetnames

print(sn)

sheet1 = wb[sn[0]]
cellA1 = sheet1['A1']

print(sheet1)
print(cellA1)

print(cellA1.value)
print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

# First coordinate is row, second is column
# Important to use numbers so one can iterate
print(sheet1.cell(1,2).value)

# Returns max column where data is
print(sheet1.max_row)
print(sheet1.max_column)

# Print Names of fruits from column B
for i in range(1, sheet1.max_row+1):
    print(sheet1.cell(i,2).value)

# To convert numbers to letter columns

print(xl.utils.get_column_letter(900))
print(xl.utils.column_index_from_string('AHP'))

# Loop thru specific part of WS
for currentrow in sheet1['A1':'C3']:
    print(currentrow)
    for currentcell in currentrow:
        print(currentcell.coordinate, currentcell.value)


for currentrow in sheet1.iter_rows(min_row = 1, max_row = sheet1.max_row):
    #print(currentrow)
    print(currentrow[0].value)
    print(currentrow[1].value)
    print(currentrow[2].value)
    
