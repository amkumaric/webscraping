import openpyxl as xl
from openpyxl.styles import Font

# Create new excel doc
wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index = 1, title = 'Second Sheet')

# Write content to a cell
ws['A1'] = 'Invoice'

headerfont= Font(name='Times New Roman', size = 24, bold = True)

ws['A1'].font = headerfont

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

# Merge
ws.merge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16,bold=True)

ws['B8'] = '=SUM(B2:B4)'

ws.column_dimensions['A'].width = 25

wb.save('PythontoExcel.xlsx')
# Writing to sheet 2

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

#for row in range(1, read_ws.max_row + 1):
    #for col in range(1,5):
        # Read value from sheet
        #cell_value = read_ws.cell(row = row, column = col).value

        # Write value to the target sheet
        #write_sheet.cell(row=row, column=col).value = cell_value

for row in read_ws.iter_rows():
    ls = [i.value for i in row]
    write_sheet.append(ls)

max_row = write_sheet.max_row

write_sheet.cell(max_row+2, 2).value = 'Total'
write_sheet.cell(max_row+2, 2).font = Font(size=16,bold=True)

write_sheet.cell(max_row+2, 3).value = '=SUM(C2:C' + str(max_row) + ')'
write_sheet.cell(max_row+2, 4).value = '=SUM(D2:C' + str(max_row) + ')'

write_sheet.cell(max_row+4, 2).value = 'Average'
write_sheet.cell(max_row+4, 2).font = Font(size=16,bold=True)

write_sheet.cell(max_row+4, 3).value = '=Average(C2:C' + str(max_row) + ')'
write_sheet.cell(max_row+4, 4).value = '=Average(D2:C' + str(max_row) + ')'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 15
write_sheet.column_dimensions['C'].width = 15
write_sheet.column_dimensions['D'].width = 15

for cell in write_sheet['C:C']:
    cell.number_format = '#,##0'

for cell in write_sheet['D:D']:
    cell.number_format = u'"$ "#,##0.00'


wb.save('PythontoExcel.xlsx')


